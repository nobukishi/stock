import os
import sqlite3
import openpyxl

wb = openpyxl.load_workbook('./automated_trading_result.xlsx')
ws = wb.active

# workbook = openpyxl.Workbook()
# sheet = workbook.active
# sheet['A1'].value = '年月日'
# sheet['B1'].value = '平均気温'
# sheet['C1'].value = '最高気温'
# sheet['D1'].value = '最低気温'

# db = sqlite3.connect('example.db')
# c = db.cursor()
# c.execute('SELECT * FROM temperature')
# for i, row in enumerate (c):
#     sheet['A' + str(i+2)].value = row[0] #年月日
#     sheet['B' + str(i+2)].value = row[1] #平均気温
#     sheet['C' + str(i+2)].value = row[2] #最高気温
#     sheet['D' + str(i+2)].value = row[3] #最低気温




db_filename = 'stock.db'
db_is_new = not os.path.exists(db_filename)
schema="""
create table buy_stock (
    id            integer primary key autoincrement not null,
    stock_code    text, 
    stock_name    text,
    buy_day       date, 
    buy_price     integer
);
"""
with sqlite3.connect(db_filename) as conn:
    if db_is_new:
        #print ('Creating schema')
        conn.execute(schema)
        conn.commit()
        conn.close()   
        
    cursor = conn.cursor()
    # cursor.execute("""
    #                 INSERT INTO buy_stock(stock_code,stock_name,buy_day,buy_price) values('8609','name','2002/1/1',2000)
                   
    #                """)
    cursor.execute("""
    select id, stock_code, stock_name, buy_day, buy_price from buy_stock 
    """)
    
    # for row in cursor.fetchall():
    #     id,stock_code, stock_name, buy_day, buy_price=row
    #     print(id,stock_code, stock_name, buy_day, buy_price)
       
    
    for i ,row in enumerate(cursor):
        print(i,row)
        ws['A' + str(i+2)].value = row[0] #id
        ws['B' + str(i+2)].value = row[1] #stock_code
        ws['C' + str(i+2)].value = row[2] #stock_name
        ws['D' + str(i+2)].value = row[3] #buy_day
        ws['E' + str(i+2)].value = row[4] #buy_price
     
wb.save('./automated_trading_result.xlsx')

        